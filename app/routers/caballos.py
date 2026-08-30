from io import BytesIO

from fastapi import (
    APIRouter,
    Depends,
    File,
    Form,
    HTTPException,
    Query,
    Request,
    UploadFile,
)
from fastapi.responses import HTMLResponse, RedirectResponse, StreamingResponse
from fastapi.templating import Jinja2Templates
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font
from sqlalchemy import or_, select
from sqlalchemy.orm import Session

from app.core.database import get_db
from app.models.caballo import Caballo
from app.models.caballo_fecha import CaballoFecha
from app.models.caballo_historial import CaballoHistorial
from app.models.categoria import Categoria
from app.models.fecha import Fecha
from app.models.tropilla import Tropilla


templates = Jinja2Templates(directory="app/templates")

router = APIRouter(
    prefix="/caballos",
    tags=["Caballos"],
)


def obtener_caballo_o_404(caballo_id: int, db: Session) -> Caballo:
    caballo = db.get(Caballo, caballo_id)

    if caballo is None:
        raise HTTPException(status_code=404, detail="Caballo no encontrado")

    return caballo


def validar_fecha_categoria(
    fecha_id: int,
    categoria_id: int,
    db: Session,
) -> tuple[Fecha, Categoria]:
    fecha = db.get(Fecha, fecha_id)
    categoria = db.get(Categoria, categoria_id)

    if fecha is None:
        raise HTTPException(status_code=404, detail="Fecha no encontrada")

    if categoria is None:
        raise HTTPException(status_code=404, detail="Categoría no encontrada")

    if categoria.campeonato_id != fecha.campeonato_id:
        raise HTTPException(
            status_code=400,
            detail="La categoría no pertenece al campeonato de la fecha seleccionada.",
        )

    return fecha, categoria


def obtener_asignacion_vigente(
    caballo_id: int,
    db: Session,
) -> CaballoFecha | None:
    return db.scalar(
        select(CaballoFecha).where(
            CaballoFecha.caballo_id == caballo_id
        )
    )


def guardar_historial_asignacion(
    db: Session,
    caballo: Caballo,
    asignacion: CaballoFecha,
    evento: str,
    observaciones: str | None = None,
):
    fecha = db.get(Fecha, asignacion.fecha_id)

    db.add(
        CaballoHistorial(
            caballo_id=caballo.id,
            campeonato_id=(fecha.campeonato_id if fecha else None),
            fecha_id=asignacion.fecha_id,
            categoria_id=asignacion.categoria_id,
            evento=evento,
            estado_caballo=caballo.estado,
            observaciones=observaciones,
        )
    )


def guardar_historial_estado(
    db: Session,
    caballo: Caballo,
    asignacion: CaballoFecha | None,
    estado_anterior: str,
    estado_nuevo: str,
):
    fecha = db.get(Fecha, asignacion.fecha_id) if asignacion else None

    db.add(
        CaballoHistorial(
            caballo_id=caballo.id,
            campeonato_id=(fecha.campeonato_id if fecha else None),
            fecha_id=(asignacion.fecha_id if asignacion else None),
            categoria_id=(asignacion.categoria_id if asignacion else None),
            evento="cambio_estado",
            estado_caballo=estado_nuevo,
            observaciones=(
                f"Estado anterior: {estado_anterior}. "
                f"Nuevo estado: {estado_nuevo}."
            ),
        )
    )


@router.get("", response_class=HTMLResponse)
def listar_caballos(
    request: Request,
    buscar: str = Query(default=""),
    estado: str = Query(default="todos"),
    fecha_id: int = Query(default=0),
    categoria_id: int = Query(default=0),
    db: Session = Depends(get_db),
):
    consulta = select(Caballo)
    texto = buscar.strip()

    if texto:
        patron = f"%{texto}%"
        consulta = (
            consulta
            .join(Tropilla, Caballo.tropilla_id == Tropilla.id, isouter=True)
            .where(
                or_(
                    Caballo.nombre.ilike(patron),
                    Caballo.pelaje.ilike(patron),
                    Tropilla.nombre.ilike(patron),
                )
            )
        )

    if estado != "todos":
        consulta = consulta.where(Caballo.estado == estado)

    if fecha_id == -1:
        consulta = consulta.where(
            Caballo.id.not_in(select(CaballoFecha.caballo_id))
        )
    elif fecha_id > 0:
        subconsulta = select(CaballoFecha.caballo_id).where(
            CaballoFecha.fecha_id == fecha_id
        )

        if categoria_id > 0:
            subconsulta = subconsulta.where(
                CaballoFecha.categoria_id == categoria_id
            )

        consulta = consulta.where(Caballo.id.in_(subconsulta))
    elif categoria_id > 0:
        consulta = consulta.where(
            Caballo.id.in_(
                select(CaballoFecha.caballo_id).where(
                    CaballoFecha.categoria_id == categoria_id
                )
            )
        )

    caballos = db.scalars(
        consulta.order_by(Caballo.nombre.asc())
    ).unique().all()

    fechas = db.scalars(
        select(Fecha).order_by(Fecha.fecha.asc())
    ).all()

    categorias = db.scalars(
        select(Categoria).order_by(
            Categoria.orden.asc(),
            Categoria.nombre.asc(),
        )
    ).all()

    asignaciones = db.scalars(select(CaballoFecha)).all()
    asignacion_por_caballo = {
        asignacion.caballo_id: asignacion
        for asignacion in asignaciones
    }

    return templates.TemplateResponse(
        request=request,
        name="caballos/listado.html",
        context={
            "caballos": caballos,
            "buscar": buscar,
            "estado": estado,
            "fecha_id": fecha_id,
            "categoria_id": categoria_id,
            "fechas": fechas,
            "categorias": categorias,
            "asignacion_por_caballo": asignacion_por_caballo,
            "importados": request.query_params.get("importados"),
            "existentes": request.query_params.get("existentes"),
            "ya_asignados": request.query_params.get("ya_asignados"),
            "reasignados_auto": request.query_params.get("reasignados_auto"),
            "errores_importacion": request.query_params.get("errores_importacion"),
            "menu_activo": "caballos",
            "usuario_nombre": request.session.get("usuario_nombre", "Administrador"),
        },
    )


@router.get("/nuevo", response_class=HTMLResponse)
def formulario_nuevo_caballo(
    request: Request,
    db: Session = Depends(get_db),
):
    tropillas = db.scalars(
        select(Tropilla)
        .where(Tropilla.activo.is_(True))
        .order_by(Tropilla.nombre.asc())
    ).all()

    return templates.TemplateResponse(
        request=request,
        name="caballos/formulario.html",
        context={
            "caballo": None,
            "tropillas": tropillas,
            "asignacion": None,
            "historial": [],
            "fechas": [],
            "categorias": [],
            "menu_activo": "caballos",
            "usuario_nombre": request.session.get("usuario_nombre", "Administrador"),
        },
    )


@router.post("/nuevo")
def crear_caballo(
    nombre: str = Form(...),
    tropilla_id: int | None = Form(None),
    pelaje: str = Form(""),
    estado: str = Form("activo"),
    observaciones: str = Form(""),
    db: Session = Depends(get_db),
):
    caballo = Caballo(
        nombre=nombre.strip(),
        tropilla_id=tropilla_id,
        pelaje=pelaje.strip() or None,
        estado=estado,
        observaciones=observaciones.strip() or None,
    )

    db.add(caballo)
    db.commit()

    return RedirectResponse(url="/caballos", status_code=303)


@router.get("/modelo-excel")
def descargar_modelo_excel():
    libro = Workbook()
    hoja = libro.active
    hoja.title = "Caballos"

    columnas = [
        "Nombre",
        "Tropilla",
        "Propietario_Tropilla",
        "Localidad_Tropilla",
        "Pelaje",
        "Estado",
        "Observaciones",
    ]

    hoja.append(columnas)

    for celda in hoja[1]:
        celda.font = Font(bold=True)

    hoja.append([
        "Retobado",
        "Los Rebeldes",
        "Mono",
        "Aguada de Guerra",
        "Trigueño oscuro",
        "activo",
        "",
    ])

    for columna, ancho in {
        "A": 25,
        "B": 25,
        "C": 25,
        "D": 25,
        "E": 22,
        "F": 15,
        "G": 40,
    }.items():
        hoja.column_dimensions[columna].width = ancho

    archivo = BytesIO()
    libro.save(archivo)
    archivo.seek(0)

    return StreamingResponse(
        archivo,
        media_type=(
            "application/vnd.openxmlformats-officedocument."
            "spreadsheetml.sheet"
        ),
        headers={
            "Content-Disposition": 'attachment; filename="modelo_caballos.xlsx"'
        },
    )


@router.post("/importar-excel")
async def importar_caballos_excel(
    request: Request,
    fecha_id: int = Form(...),
    categoria_id: int = Form(...),
    archivo: UploadFile = File(...),
    db: Session = Depends(get_db),
):
    fecha_destino, _ = validar_fecha_categoria(
        fecha_id,
        categoria_id,
        db,
    )

    if not archivo.filename:
        raise HTTPException(status_code=400, detail="Debe seleccionar un archivo.")

    if not archivo.filename.lower().endswith(".xlsx"):
        raise HTTPException(status_code=400, detail="El archivo debe ser XLSX.")

    contenido = await archivo.read()

    try:
        libro = load_workbook(BytesIO(contenido), data_only=True)
    except Exception as exc:
        raise HTTPException(
            status_code=400,
            detail="No se pudo leer el archivo Excel.",
        ) from exc

    hoja = libro.active

    encabezados = [
        str(celda.value).strip() if celda.value is not None else ""
        for celda in hoja[1]
    ]

    columnas_esperadas = [
        "Nombre",
        "Tropilla",
        "Propietario_Tropilla",
        "Localidad_Tropilla",
        "Pelaje",
        "Estado",
        "Observaciones",
    ]

    if encabezados[:7] != columnas_esperadas:
        raise HTTPException(
            status_code=400,
            detail="El formato del Excel no corresponde al modelo.",
        )

    importados = 0
    existentes = 0
    ya_asignados = 0
    reasignados_auto = 0
    errores = 0
    conflictos: list[int] = []

    for fila in hoja.iter_rows(min_row=2, values_only=True):
        valores = list(fila[:7])

        if not any(
            valor is not None and str(valor).strip()
            for valor in valores
        ):
            continue

        (
            nombre,
            nombre_tropilla,
            propietario,
            localidad,
            pelaje,
            estado_excel,
            observaciones,
        ) = valores

        nombre = str(nombre).strip() if nombre is not None else ""
        nombre_tropilla = (
            str(nombre_tropilla).strip()
            if nombre_tropilla is not None
            else ""
        )

        if nombre == "-":
            nombre = ""
        if nombre_tropilla == "-":
            nombre_tropilla = ""

        if not nombre or not nombre_tropilla:
            errores += 1
            continue

        tropilla = db.scalar(
            select(Tropilla).where(
                Tropilla.nombre.ilike(nombre_tropilla)
            )
        )

        if tropilla is None:
            propietario_limpio = (
                str(propietario).strip()
                if propietario is not None
                else ""
            )
            localidad_limpia = (
                str(localidad).strip()
                if localidad is not None
                else ""
            )

            tropilla = Tropilla(
                nombre=nombre_tropilla,
                propietario=(None if propietario_limpio in {"", "-"} else propietario_limpio),
                localidad=(None if localidad_limpia in {"", "-"} else localidad_limpia),
                provincia="Rio Negro",
                activo=True,
            )

            db.add(tropilla)
            db.flush()

        caballo = db.scalar(
            select(Caballo).where(
                Caballo.nombre.ilike(nombre),
                Caballo.tropilla_id == tropilla.id,
            )
        )

        if caballo is None:
            pelaje_limpio = str(pelaje).strip() if pelaje is not None else ""
            observaciones_limpias = (
                str(observaciones).strip()
                if observaciones is not None
                else ""
            )
            estado_limpio = (
                str(estado_excel).strip().lower()
                if estado_excel is not None
                else "activo"
            )

            if not estado_limpio or estado_limpio == "-":
                estado_limpio = "activo"

            if estado_limpio not in {
                "activo",
                "inactivo",
                "lesionado",
                "retirado",
            }:
                estado_limpio = "activo"

            caballo = Caballo(
                nombre=nombre,
                tropilla_id=tropilla.id,
                pelaje=(None if pelaje_limpio in {"", "-"} else pelaje_limpio),
                estado=estado_limpio,
                observaciones=(
                    None
                    if observaciones_limpias in {"", "-"}
                    else observaciones_limpias
                ),
            )

            db.add(caballo)
            db.flush()
            importados += 1
        else:
            existentes += 1

        asignacion = obtener_asignacion_vigente(caballo.id, db)

        if asignacion is None:
            db.add(
                CaballoFecha(
                    caballo_id=caballo.id,
                    fecha_id=fecha_id,
                    categoria_id=categoria_id,
                )
            )
            continue

        if (
            asignacion.fecha_id == fecha_id
            and asignacion.categoria_id == categoria_id
        ):
            ya_asignados += 1
            continue

        fecha_actual = db.get(Fecha, asignacion.fecha_id)

        # Reasignación automática solo si viene de OTRA fecha ya sorteada.
        if (
            fecha_actual is not None
            and fecha_actual.sorteada
            and asignacion.fecha_id != fecha_id
        ):
            guardar_historial_asignacion(
                db,
                caballo,
                asignacion,
                evento="participacion_archivada",
                observaciones=(
                    "Reasignación automática por importación "
                    f"a {fecha_destino.nombre}."
                ),
            )

            asignacion.fecha_id = fecha_id
            asignacion.categoria_id = categoria_id
            reasignados_auto += 1
            continue

        conflictos.append(caballo.id)

    db.commit()

    if conflictos:
        request.session["importacion_caballos_conflictos"] = {
            "fecha_destino_id": fecha_id,
            "categoria_destino_id": categoria_id,
            "caballo_ids": conflictos,
        }

        return RedirectResponse(
            url="/caballos/importar-conflictos",
            status_code=303,
        )

    return RedirectResponse(
        url=(
            "/caballos"
            f"?importados={importados}"
            f"&existentes={existentes}"
            f"&ya_asignados={ya_asignados}"
            f"&reasignados_auto={reasignados_auto}"
            f"&errores_importacion={errores}"
        ),
        status_code=303,
    )


@router.get("/importar-conflictos", response_class=HTMLResponse)
def mostrar_conflictos_importacion(
    request: Request,
    db: Session = Depends(get_db),
):
    pendiente = request.session.get("importacion_caballos_conflictos")

    if not pendiente:
        return RedirectResponse(url="/caballos", status_code=303)

    fecha_destino = db.get(Fecha, pendiente["fecha_destino_id"])
    categoria_destino = db.get(Categoria, pendiente["categoria_destino_id"])

    conflictos = []

    for caballo_id in pendiente["caballo_ids"]:
        caballo = db.get(Caballo, caballo_id)
        asignacion = obtener_asignacion_vigente(caballo_id, db)

        if caballo is None or asignacion is None:
            continue

        conflictos.append({
            "caballo": caballo,
            "asignacion": asignacion,
        })

    return templates.TemplateResponse(
        request=request,
        name="caballos/importar_conflictos.html",
        context={
            "conflictos": conflictos,
            "fecha_destino": fecha_destino,
            "categoria_destino": categoria_destino,
            "menu_activo": "caballos",
            "usuario_nombre": request.session.get("usuario_nombre", "Administrador"),
        },
    )


@router.post("/importar-conflictos/reasignar")
def confirmar_reasignacion_importacion(
    request: Request,
    db: Session = Depends(get_db),
):
    pendiente = request.session.get("importacion_caballos_conflictos")

    if not pendiente:
        return RedirectResponse(url="/caballos", status_code=303)

    fecha_id = pendiente["fecha_destino_id"]
    categoria_id = pendiente["categoria_destino_id"]

    fecha_destino, _ = validar_fecha_categoria(
        fecha_id,
        categoria_id,
        db,
    )

    reasignados = 0

    for caballo_id in pendiente["caballo_ids"]:
        caballo = db.get(Caballo, caballo_id)
        asignacion = obtener_asignacion_vigente(caballo_id, db)

        if caballo is None or asignacion is None:
            continue

        if (
            asignacion.fecha_id == fecha_id
            and asignacion.categoria_id == categoria_id
        ):
            continue

        guardar_historial_asignacion(
            db,
            caballo,
            asignacion,
            evento="reasignado_confirmado",
            observaciones=(
                "Reasignación confirmada por el usuario "
                f"a {fecha_destino.nombre}."
            ),
        )

        asignacion.fecha_id = fecha_id
        asignacion.categoria_id = categoria_id
        reasignados += 1

    db.commit()
    request.session.pop("importacion_caballos_conflictos", None)

    return RedirectResponse(
        url=f"/caballos?reasignados_auto={reasignados}",
        status_code=303,
    )


@router.post("/importar-conflictos/cancelar")
def cancelar_reasignacion_importacion(request: Request):
    request.session.pop("importacion_caballos_conflictos", None)
    return RedirectResponse(url="/caballos", status_code=303)


@router.get("/{caballo_id}/editar", response_class=HTMLResponse)
def formulario_editar_caballo(
    caballo_id: int,
    request: Request,
    db: Session = Depends(get_db),
):
    caballo = obtener_caballo_o_404(caballo_id, db)

    tropillas = db.scalars(
        select(Tropilla).order_by(Tropilla.nombre.asc())
    ).all()

    fechas = db.scalars(
        select(Fecha).order_by(Fecha.fecha.asc())
    ).all()

    categorias = db.scalars(
        select(Categoria).order_by(
            Categoria.orden.asc(),
            Categoria.nombre.asc(),
        )
    ).all()

    asignacion = obtener_asignacion_vigente(caballo_id, db)

    historial = db.scalars(
        select(CaballoHistorial)
        .where(CaballoHistorial.caballo_id == caballo_id)
        .order_by(CaballoHistorial.creado_en.desc())
    ).all()

    return templates.TemplateResponse(
        request=request,
        name="caballos/formulario.html",
        context={
            "caballo": caballo,
            "tropillas": tropillas,
            "fechas": fechas,
            "categorias": categorias,
            "asignacion": asignacion,
            "historial": historial,
            "menu_activo": "caballos",
            "usuario_nombre": request.session.get("usuario_nombre", "Administrador"),
        },
    )


@router.post("/{caballo_id}/asignaciones/{asignacion_id}/dejar-libre")
def dejar_libre_caballo(
    caballo_id: int,
    asignacion_id: int,
    db: Session = Depends(get_db),
):
    caballo = obtener_caballo_o_404(caballo_id, db)
    asignacion = db.get(CaballoFecha, asignacion_id)

    if asignacion is None or asignacion.caballo_id != caballo.id:
        raise HTTPException(status_code=404, detail="Asignación no encontrada")

    guardar_historial_asignacion(
        db,
        caballo,
        asignacion,
        evento="liberado",
        observaciones="El caballo fue dejado libre manualmente.",
    )

    db.delete(asignacion)
    db.commit()

    return RedirectResponse(
        url=f"/caballos/{caballo_id}/editar",
        status_code=303,
    )


@router.post("/{caballo_id}/editar")
def editar_caballo(
    caballo_id: int,
    nombre: str = Form(...),
    tropilla_id: int | None = Form(None),
    pelaje: str = Form(""),
    estado: str = Form("activo"),
    observaciones: str = Form(""),
    fecha_id: int | None = Form(None),
    categoria_id: int | None = Form(None),
    db: Session = Depends(get_db),
):
    caballo = obtener_caballo_o_404(caballo_id, db)
    asignacion = obtener_asignacion_vigente(caballo_id, db)
    estado_anterior = caballo.estado

    if estado != estado_anterior:
        guardar_historial_estado(
            db,
            caballo,
            asignacion,
            estado_anterior,
            estado,
        )

    caballo.nombre = nombre.strip()
    caballo.tropilla_id = tropilla_id
    caballo.pelaje = pelaje.strip() or None
    caballo.estado = estado
    caballo.observaciones = observaciones.strip() or None

    if fecha_id or categoria_id:
        if not fecha_id or not categoria_id:
            raise HTTPException(
                status_code=400,
                detail="Debe seleccionar Fecha y Categoría.",
            )

        validar_fecha_categoria(fecha_id, categoria_id, db)

        if asignacion is None:
            db.add(
                CaballoFecha(
                    caballo_id=caballo.id,
                    fecha_id=fecha_id,
                    categoria_id=categoria_id,
                )
            )
        elif (
            asignacion.fecha_id != fecha_id
            or asignacion.categoria_id != categoria_id
        ):
            guardar_historial_asignacion(
                db,
                caballo,
                asignacion,
                evento="reasignado_manual",
                observaciones=(
                    "Asignación modificada desde la ficha individual del caballo."
                ),
            )

            asignacion.fecha_id = fecha_id
            asignacion.categoria_id = categoria_id

    db.commit()
    return RedirectResponse(url="/caballos", status_code=303)
