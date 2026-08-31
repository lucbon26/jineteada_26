
from datetime import date, datetime
from io import BytesIO

from fastapi import (
    APIRouter,
    Depends,
    File,
    Form,
    HTTPException,
    Query,
    Request,
    UploadFile,
)
from fastapi.responses import HTMLResponse, RedirectResponse, StreamingResponse
from fastapi.templating import Jinja2Templates
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font
from sqlalchemy import or_, select
from sqlalchemy.orm import Session

from app.core.database import get_db
from app.models.campeonato import Campeonato
from app.models.categoria import Categoria
from app.models.fecha import Fecha
from app.models.jinete import Jinete
from app.models.jinete_campeonato import JineteCampeonato
from app.models.jinete_fecha import JineteFecha


templates = Jinja2Templates(directory="app/templates")

router = APIRouter(
    prefix="/jinetes",
    tags=["Jinetes"],
)


def limpiar(valor) -> str:
    if valor is None:
        return ""
    texto = str(valor).strip()
    return "" if texto == "-" else texto


def dni_limpio(valor) -> str:
    texto = limpiar(valor)
    if texto.endswith(".0"):
        texto = texto[:-2]
    return "".join(ch for ch in texto if ch.isdigit())


def normalizar_identidad(valor: str) -> str:
    return " ".join(
        limpiar(valor).upper().split()
    )


def parsear_fecha_excel(valor) -> date | None:
    if valor in (None, "", "-"):
        return None

    if isinstance(valor, datetime):
        return valor.date()

    if isinstance(valor, date):
        return valor

    texto = str(valor).strip()

    for formato in ("%d/%m/%Y", "%Y-%m-%d", "%d-%m-%Y"):
        try:
            return datetime.strptime(texto, formato).date()
        except ValueError:
            pass

    return None


def obtener_jinete_o_404(jinete_id: int, db: Session) -> Jinete:
    jinete = db.get(Jinete, jinete_id)
    if jinete is None:
        raise HTTPException(status_code=404, detail="Jinete no encontrado")
    return jinete


def obtener_asignacion_campeonato(
    jinete_id: int,
    campeonato_id: int,
    db: Session,
) -> JineteCampeonato | None:
    return db.scalar(
        select(JineteCampeonato).where(
            JineteCampeonato.jinete_id == jinete_id,
            JineteCampeonato.campeonato_id == campeonato_id,
        )
    )


def obtener_participacion_fecha(
    jinete_id: int,
    fecha_id: int,
    db: Session,
) -> JineteFecha | None:
    return db.scalar(
        select(JineteFecha).where(
            JineteFecha.jinete_id == jinete_id,
            JineteFecha.fecha_id == fecha_id,
        )
    )


def obtener_categoria_por_nombre(
    campeonato_id: int,
    nombre: str,
    db: Session,
) -> Categoria | None:
    nombre = nombre.strip()
    if not nombre:
        return None

    return db.scalar(
        select(Categoria).where(
            Categoria.campeonato_id == campeonato_id,
            Categoria.nombre.ilike(nombre),
        )
    )


@router.get("", response_class=HTMLResponse)
def listar_jinetes(
    request: Request,
    buscar: str = Query(default=""),
    estado: str = Query(default="todos"),
    fecha_id: int = Query(default=0),
    categoria_id: int = Query(default=0),
    db: Session = Depends(get_db),
):
    consulta = select(Jinete)
    texto = buscar.strip()

    if texto:
        patron = f"%{texto}%"
        consulta = consulta.where(
            or_(
                Jinete.nombres.ilike(patron),
                Jinete.apellidos.ilike(patron),
                Jinete.dni.ilike(patron),
                Jinete.localidad.ilike(patron),
            )
        )

    if estado != "todos":
        consulta = consulta.where(Jinete.estado == estado)

    if fecha_id == -1:
        consulta = consulta.where(
            Jinete.id.not_in(select(JineteFecha.jinete_id))
        )
    elif fecha_id > 0:
        subconsulta = select(JineteFecha.jinete_id).where(
            JineteFecha.fecha_id == fecha_id
        )

        if categoria_id > 0:
            subconsulta = subconsulta.where(
                JineteFecha.categoria_id == categoria_id
            )

        consulta = consulta.where(
            Jinete.id.in_(subconsulta)
        )
    elif categoria_id > 0:
        consulta = consulta.where(
            Jinete.id.in_(
                select(JineteCampeonato.jinete_id).where(
                    JineteCampeonato.categoria_id == categoria_id
                )
            )
        )

    jinetes = db.scalars(
        consulta.order_by(
            Jinete.apellidos.asc(),
            Jinete.nombres.asc(),
        )
    ).unique().all()

    fechas = db.scalars(
        select(Fecha).order_by(Fecha.fecha.asc())
    ).all()

    categorias = db.scalars(
        select(Categoria).order_by(
            Categoria.orden.asc(),
            Categoria.nombre.asc(),
        )
    ).all()

    asignaciones_campeonato = db.scalars(
        select(JineteCampeonato)
        .where(JineteCampeonato.categoria_id.is_not(None))
        .order_by(JineteCampeonato.id.desc())
    ).all()

    categorias_por_jinete = {}

    for asignacion in asignaciones_campeonato:
        categoria = db.get(Categoria, asignacion.categoria_id)

        if categoria is None:
            continue

        categorias_por_jinete.setdefault(
            asignacion.jinete_id,
            [],
        )

        if categoria.nombre not in categorias_por_jinete[asignacion.jinete_id]:
            categorias_por_jinete[asignacion.jinete_id].append(
                categoria.nombre
            )

    return templates.TemplateResponse(
        request=request,
        name="jinetes/listado.html",
        context={
            "jinetes": jinetes,
            "buscar": buscar,
            "estado": estado,
            "fecha_id": fecha_id,
            "categoria_id": categoria_id,
            "fechas": fechas,
            "categorias": categorias,
            "categorias_por_jinete": categorias_por_jinete,
            "importados": request.query_params.get("importados"),
            "existentes": request.query_params.get("existentes"),
            "ya_en_fecha": request.query_params.get("ya_en_fecha"),
            "avanzados": request.query_params.get("avanzados"),
            "errores_importacion": request.query_params.get("errores_importacion"),
            "menu_activo": "jinetes",
            "usuario_nombre": request.session.get(
                "usuario_nombre",
                "Administrador",
            ),
        },
    )


@router.get("/nuevo", response_class=HTMLResponse)
def formulario_nuevo_jinete(request: Request):
    return templates.TemplateResponse(
        request=request,
        name="jinetes/formulario.html",
        context={
            "accion": "Nuevo",
            "jinete": None,
            "menu_activo": "jinetes",
            "usuario_nombre": request.session.get("usuario_nombre", "Administrador"),
        },
    )


@router.post("/nuevo")
def crear_jinete(
    nombres: str = Form(...),
    apellidos: str = Form(...),
    dni: str = Form(...),
    fecha_nacimiento: date | None = Form(None),
    sexo: str = Form(""),
    nacionalidad: str = Form(""),
    celular: str = Form(""),
    email: str = Form(""),
    provincia: str = Form(""),
    localidad: str = Form(""),
    categoria_habitual: str = Form(""),
    club_agrupacion: str = Form(""),
    observaciones: str = Form(""),
    estado: str = Form("activo"),
    db: Session = Depends(get_db),
):
    dni_normalizado = dni_limpio(dni)

    if not dni_normalizado:
        raise HTTPException(status_code=400, detail="Debe ingresar un DNI válido.")

    existente = db.scalar(
        select(Jinete).where(Jinete.dni == dni_normalizado)
    )

    if existente:
        raise HTTPException(
            status_code=400,
            detail="Ya existe un jinete registrado con ese DNI",
        )

    nuevo = Jinete(
        nombres=nombres.strip(),
        apellidos=apellidos.strip(),
        dni=dni_normalizado,
        fecha_nacimiento=fecha_nacimiento,
        sexo=sexo.strip() or None,
        nacionalidad=nacionalidad.strip() or None,
        celular=celular.strip() or None,
        email=email.strip() or None,
        provincia=provincia.strip() or None,
        localidad=localidad.strip() or None,
        categoria_habitual=categoria_habitual.strip() or None,
        club_agrupacion=club_agrupacion.strip() or None,
        observaciones=observaciones.strip() or None,
        estado=estado,
    )

    db.add(nuevo)
    db.commit()

    return RedirectResponse(url="/jinetes", status_code=303)


@router.get("/modelo-excel")
def descargar_modelo_excel():
    libro = Workbook()
    hoja = libro.active
    hoja.title = "Jinetes"

    columnas = [
        "Apellidos",
        "Nombres",
        "DNI",
        "Fecha_Nacimiento",
        "Celular",
        "Localidad",
    ]

    hoja.append(columnas)

    for celda in hoja[1]:
        celda.font = Font(bold=True)

    hoja.append([
        "Morales",
        "Sebastian",
        "40111075",
        "15/02/1997",
        "2995790409",
        "Catriel",
    ])

    for columna, ancho in {
        "A": 24,
        "B": 24,
        "C": 16,
        "D": 20,
        "E": 18,
        "F": 24,
    }.items():
        hoja.column_dimensions[columna].width = ancho

    archivo = BytesIO()
    libro.save(archivo)
    archivo.seek(0)

    return StreamingResponse(
        archivo,
        media_type=(
            "application/vnd.openxmlformats-officedocument."
            "spreadsheetml.sheet"
        ),
        headers={
            "Content-Disposition": 'attachment; filename="modelo_jinetes.xlsx"'
        },
    )


@router.post("/importar-excel")
async def importar_jinetes_excel(
    request: Request,
    fecha_id: int = Form(...),
    categoria_id: int = Form(...),
    archivo: UploadFile = File(...),
    db: Session = Depends(get_db),
):
    fecha = db.get(Fecha, fecha_id)
    categoria_destino = db.get(Categoria, categoria_id)

    if fecha is None:
        raise HTTPException(status_code=404, detail="Fecha no encontrada.")

    if categoria_destino is None:
        raise HTTPException(status_code=404, detail="Categoría no encontrada.")

    if categoria_destino.campeonato_id != fecha.campeonato_id:
        raise HTTPException(
            status_code=400,
            detail="La categoría no pertenece al campeonato de la fecha seleccionada.",
        )

    campeonato_id = fecha.campeonato_id

    if not archivo.filename:
        raise HTTPException(
            status_code=400,
            detail="Debe seleccionar un archivo.",
        )

    if not archivo.filename.lower().endswith(".xlsx"):
        raise HTTPException(
            status_code=400,
            detail="El archivo debe ser XLSX.",
        )

    contenido = await archivo.read()

    try:
        libro = load_workbook(BytesIO(contenido), data_only=True)
    except Exception as exc:
        raise HTTPException(
            status_code=400,
            detail="No se pudo leer el archivo Excel.",
        ) from exc

    hoja = libro.active

    encabezados = [
        str(celda.value).strip() if celda.value is not None else ""
        for celda in hoja[1]
    ]

    columnas_modelo_actual = [
        "Apellidos",
        "Nombres",
        "DNI",
        "Fecha_Nacimiento",
        "Celular",
        "Localidad",
    ]

    columnas_modelo_anterior = [
        "Nombres",
        "Apellidos",
        "DNI",
        "Fecha_Nacimiento",
        "Celular",
        "Localidad",
    ]

    usa_modelo_actual = encabezados[:6] == columnas_modelo_actual
    usa_modelo_anterior = encabezados[:6] == columnas_modelo_anterior

    if not usa_modelo_actual and not usa_modelo_anterior:
        raise HTTPException(
            status_code=400,
            detail=(
                "El formato del Excel no corresponde al modelo de Jinetes. "
                "Descargue nuevamente el modelo desde el sistema."
            ),
        )

    importados = 0
    existentes = 0
    ya_en_fecha = 0
    avanzados = 0
    errores = 0
    pendientes = []
    dni_conflictos = []

    for numero_fila, fila in enumerate(
        hoja.iter_rows(min_row=2, values_only=True),
        start=2,
    ):
        valores = list(fila[:6])

        if not any(limpiar(valor) for valor in valores):
            continue

        if usa_modelo_actual:
            (
                apellidos,
                nombres,
                dni,
                fecha_nacimiento,
                celular,
                localidad,
            ) = valores
        else:
            (
                nombres,
                apellidos,
                dni,
                fecha_nacimiento,
                celular,
                localidad,
            ) = valores

        nombres = limpiar(nombres)
        apellidos = limpiar(apellidos)
        dni = dni_limpio(dni)
        celular = limpiar(celular)
        localidad = limpiar(localidad)
        nacimiento = parsear_fecha_excel(fecha_nacimiento)

        if not nombres or not apellidos or not dni:
            errores += 1
            continue

        # Si vino una fecha escrita pero no se pudo interpretar, se informa como error.
        if limpiar(fecha_nacimiento) and nacimiento is None:
            errores += 1
            continue

        jinete = db.scalar(
            select(Jinete).where(Jinete.dni == dni)
        )

        if jinete is None:
            jinete = Jinete(
                nombres=nombres,
                apellidos=apellidos,
                dni=dni,
                fecha_nacimiento=nacimiento,
                celular=celular or None,
                localidad=localidad or None,
                estado="activo",
            )
            db.add(jinete)
            db.flush()
            importados += 1
        else:
            nombre_excel = normalizar_identidad(nombres)
            apellido_excel = normalizar_identidad(apellidos)
            nombre_base = normalizar_identidad(jinete.nombres)
            apellido_base = normalizar_identidad(jinete.apellidos)

            if (
                nombre_excel != nombre_base
                or apellido_excel != apellido_base
            ):
                dni_conflictos.append({
                    "fila": numero_fila,
                    "dni": dni,
                    "excel_apellidos": apellidos,
                    "excel_nombres": nombres,
                    "base_apellidos": jinete.apellidos,
                    "base_nombres": jinete.nombres,
                })
                continue

            existentes += 1

            if not jinete.fecha_nacimiento and nacimiento:
                jinete.fecha_nacimiento = nacimiento
            if not jinete.celular and celular:
                jinete.celular = celular
            if not jinete.localidad and localidad:
                jinete.localidad = localidad

        asignacion = obtener_asignacion_campeonato(
            jinete.id,
            campeonato_id,
            db,
        )

        if asignacion is None:
            asignacion = JineteCampeonato(
                jinete_id=jinete.id,
                campeonato_id=campeonato_id,
                categoria_id=categoria_id,
            )
            db.add(asignacion)
            db.flush()

        elif asignacion.categoria_id is None:
            asignacion.categoria_id = categoria_id

        elif asignacion.categoria_id != categoria_id:
            categoria_actual = db.get(
                Categoria,
                asignacion.categoria_id,
            )

            pendientes.append({
                "jinete_id": jinete.id,
                "fila": numero_fila,
                "motivo": "categoria_conflicto",
                "categoria_excel": categoria_destino.nombre,
                "categoria_actual": (
                    categoria_actual.nombre
                    if categoria_actual
                    else ""
                ),
            })
            continue

        participacion = obtener_participacion_fecha(
            jinete.id,
            fecha_id,
            db,
        )

        if participacion is not None:
            ya_en_fecha += 1
            continue

        db.add(
            JineteFecha(
                jinete_id=jinete.id,
                fecha_id=fecha_id,
                categoria_id=categoria_id,
            )
        )
        avanzados += 1

    db.commit()

    if dni_conflictos:
        request.session["importacion_jinetes_dni_conflictos"] = {
            "fecha_id": fecha_id,
            "categoria_destino_id": categoria_id,
            "items": dni_conflictos,
            "resumen": {
                "importados": importados,
                "existentes": existentes,
                "ya_en_fecha": ya_en_fecha,
                "avanzados": avanzados,
                "errores": errores,
            },
        }

        if pendientes:
            request.session["importacion_jinetes_pendientes"] = {
                "fecha_id": fecha_id,
                "categoria_destino_id": categoria_id,
                "items": pendientes,
            }

        return RedirectResponse(
            url="/jinetes/importar-dni-conflictos",
            status_code=303,
        )

    if pendientes:
        request.session["importacion_jinetes_pendientes"] = {
            "fecha_id": fecha_id,
            "categoria_destino_id": categoria_id,
            "items": pendientes,
        }

        return RedirectResponse(
            url="/jinetes/importar-pendientes",
            status_code=303,
        )

    return RedirectResponse(
        url=(
            "/jinetes"
            f"?importados={importados}"
            f"&existentes={existentes}"
            f"&ya_en_fecha={ya_en_fecha}"
            f"&avanzados={avanzados}"
            f"&errores_importacion={errores}"
        ),
        status_code=303,
    )


@router.get("/importar-dni-conflictos", response_class=HTMLResponse)
def mostrar_conflictos_dni(
    request: Request,
):
    pendiente = request.session.get(
        "importacion_jinetes_dni_conflictos"
    )

    if not pendiente:
        return RedirectResponse(
            url="/jinetes",
            status_code=303,
        )

    return templates.TemplateResponse(
        request=request,
        name="jinetes/importar_dni_conflictos.html",
        context={
            "items": pendiente.get("items", []),
            "resumen": pendiente.get("resumen", {}),
            "menu_activo": "jinetes",
            "usuario_nombre": request.session.get(
                "usuario_nombre",
                "Administrador",
            ),
        },
    )


@router.post("/importar-dni-conflictos/continuar")
def continuar_despues_conflictos_dni(
    request: Request,
):
    request.session.pop(
        "importacion_jinetes_dni_conflictos",
        None,
    )

    if request.session.get(
        "importacion_jinetes_pendientes"
    ):
        return RedirectResponse(
            url="/jinetes/importar-pendientes",
            status_code=303,
        )

    return RedirectResponse(
        url="/jinetes",
        status_code=303,
    )


@router.get("/importar-pendientes", response_class=HTMLResponse)
def mostrar_pendientes_importacion(
    request: Request,
    db: Session = Depends(get_db),
):
    pendiente = request.session.get("importacion_jinetes_pendientes")

    if not pendiente:
        return RedirectResponse(url="/jinetes", status_code=303)

    fecha = db.get(Fecha, pendiente["fecha_id"])
    if fecha is None:
        request.session.pop("importacion_jinetes_pendientes", None)
        return RedirectResponse(url="/jinetes", status_code=303)

    categorias = db.scalars(
        select(Categoria)
        .where(Categoria.campeonato_id == fecha.campeonato_id)
        .order_by(Categoria.orden.asc(), Categoria.nombre.asc())
    ).all()

    items = []

    for item in pendiente["items"]:
        jinete = db.get(Jinete, item["jinete_id"])
        if jinete is None:
            continue

        asignacion = obtener_asignacion_campeonato(
            jinete.id,
            fecha.campeonato_id,
            db,
        )

        categoria_actual = (
            db.get(Categoria, asignacion.categoria_id)
            if asignacion and asignacion.categoria_id
            else None
        )

        items.append({
            "jinete": jinete,
            "fila": item["fila"],
            "motivo": item["motivo"],
            "categoria_excel": item.get("categoria_excel", ""),
            "categoria_actual": categoria_actual,
        })

    return templates.TemplateResponse(
        request=request,
        name="jinetes/importar_pendientes.html",
        context={
            "fecha": fecha,
            "categorias": categorias,
            "items": items,
            "menu_activo": "jinetes",
            "usuario_nombre": request.session.get("usuario_nombre", "Administrador"),
        },
    )


@router.post("/importar-pendientes/resolver")
async def resolver_pendientes_importacion(
    request: Request,
    db: Session = Depends(get_db),
):
    pendiente = request.session.get("importacion_jinetes_pendientes")

    if not pendiente:
        return RedirectResponse(url="/jinetes", status_code=303)

    fecha = db.get(Fecha, pendiente["fecha_id"])
    if fecha is None:
        request.session.pop("importacion_jinetes_pendientes", None)
        return RedirectResponse(url="/jinetes", status_code=303)

    formulario = await request.form()
    resueltos = 0

    for item in pendiente["items"]:
        jinete_id = item["jinete_id"]
        campo = f"categoria_{jinete_id}"
        valor = formulario.get(campo)

        if not valor:
            continue

        try:
            categoria_id = int(valor)
        except (TypeError, ValueError):
            continue

        categoria = db.get(Categoria, categoria_id)

        if (
            categoria is None
            or categoria.campeonato_id != fecha.campeonato_id
        ):
            continue

        jinete = db.get(Jinete, jinete_id)
        if jinete is None:
            continue

        asignacion = obtener_asignacion_campeonato(
            jinete_id,
            fecha.campeonato_id,
            db,
        )

        if asignacion is None:
            asignacion = JineteCampeonato(
                jinete_id=jinete_id,
                campeonato_id=fecha.campeonato_id,
                categoria_id=categoria_id,
            )
            db.add(asignacion)
        else:
            asignacion.categoria_id = categoria_id

        participacion = obtener_participacion_fecha(
            jinete_id,
            fecha.id,
            db,
        )

        if participacion is None:
            db.add(
                JineteFecha(
                    jinete_id=jinete_id,
                    fecha_id=fecha.id,
                    categoria_id=categoria_id,
                )
            )
        else:
            participacion.categoria_id = categoria_id

        resueltos += 1

    db.commit()
    request.session.pop("importacion_jinetes_pendientes", None)

    return RedirectResponse(
        url=f"/jinetes?avanzados={resueltos}",
        status_code=303,
    )


@router.post("/importar-pendientes/cancelar")
def cancelar_pendientes_importacion(request: Request):
    request.session.pop("importacion_jinetes_pendientes", None)
    return RedirectResponse(url="/jinetes", status_code=303)



@router.post("/eliminar-masivo")
async def eliminar_jinetes_masivo(
    request: Request,
    db: Session = Depends(get_db),
):
    formulario = await request.form()

    ids = []
    for valor in formulario.getlist("seleccionados"):
        try:
            ids.append(int(valor))
        except (TypeError, ValueError):
            pass

    ids = list(dict.fromkeys(ids))
    eliminados = 0

    for jinete_id in ids:
        jinete = db.get(Jinete, jinete_id)

        if jinete is None:
            continue

        for participacion in db.scalars(
            select(JineteFecha).where(
                JineteFecha.jinete_id == jinete_id
            )
        ).all():
            db.delete(participacion)

        for asignacion in db.scalars(
            select(JineteCampeonato).where(
                JineteCampeonato.jinete_id == jinete_id
            )
        ).all():
            db.delete(asignacion)

        db.delete(jinete)
        eliminados += 1

    db.commit()

    return RedirectResponse(
        url=f"/jinetes?eliminados={eliminados}",
        status_code=303,
    )


@router.get("/{jinete_id}", response_class=HTMLResponse)
def detalle_jinete(
    jinete_id: int,
    request: Request,
    db: Session = Depends(get_db),
):
    jinete = obtener_jinete_o_404(jinete_id, db)

    campeonatos = db.scalars(
        select(JineteCampeonato)
        .where(JineteCampeonato.jinete_id == jinete_id)
        .order_by(JineteCampeonato.id.desc())
    ).all()

    participaciones = db.scalars(
        select(JineteFecha)
        .where(JineteFecha.jinete_id == jinete_id)
        .order_by(JineteFecha.creado_en.desc())
    ).all()

    return templates.TemplateResponse(
        request=request,
        name="jinetes/detalle.html",
        context={
            "jinete": jinete,
            "campeonatos_jinete": campeonatos,
            "participaciones": participaciones,
            "menu_activo": "jinetes",
            "usuario_nombre": request.session.get("usuario_nombre", "Administrador"),
        },
    )


@router.get("/{jinete_id}/editar", response_class=HTMLResponse)
def formulario_editar_jinete(
    jinete_id: int,
    request: Request,
    db: Session = Depends(get_db),
):
    jinete = obtener_jinete_o_404(jinete_id, db)

    return templates.TemplateResponse(
        request=request,
        name="jinetes/formulario.html",
        context={
            "accion": "Editar",
            "jinete": jinete,
            "menu_activo": "jinetes",
            "usuario_nombre": request.session.get("usuario_nombre", "Administrador"),
        },
    )


@router.post("/{jinete_id}/eliminar")
def eliminar_jinete(
    jinete_id: int,
    db: Session = Depends(get_db),
):
    jinete = obtener_jinete_o_404(jinete_id, db)

    for participacion in db.scalars(
        select(JineteFecha).where(
            JineteFecha.jinete_id == jinete_id
        )
    ).all():
        db.delete(participacion)

    for asignacion in db.scalars(
        select(JineteCampeonato).where(
            JineteCampeonato.jinete_id == jinete_id
        )
    ).all():
        db.delete(asignacion)

    db.delete(jinete)
    db.commit()

    return RedirectResponse(
        url="/jinetes",
        status_code=303,
    )


@router.post("/{jinete_id}/editar")
def editar_jinete(
    jinete_id: int,
    nombres: str = Form(...),
    apellidos: str = Form(...),
    dni: str = Form(...),
    fecha_nacimiento: date | None = Form(None),
    sexo: str = Form(""),
    nacionalidad: str = Form(""),
    celular: str = Form(""),
    email: str = Form(""),
    provincia: str = Form(""),
    localidad: str = Form(""),
    categoria_habitual: str = Form(""),
    club_agrupacion: str = Form(""),
    observaciones: str = Form(""),
    estado: str = Form("activo"),
    db: Session = Depends(get_db),
):
    jinete = obtener_jinete_o_404(jinete_id, db)
    dni_normalizado = dni_limpio(dni)

    repetido = db.scalar(
        select(Jinete).where(
            Jinete.dni == dni_normalizado,
            Jinete.id != jinete_id,
        )
    )

    if repetido:
        raise HTTPException(
            status_code=400,
            detail="Ya existe otro jinete registrado con ese DNI",
        )

    jinete.nombres = nombres.strip()
    jinete.apellidos = apellidos.strip()
    jinete.dni = dni_normalizado
    jinete.fecha_nacimiento = fecha_nacimiento
    jinete.sexo = sexo.strip() or None
    jinete.nacionalidad = nacionalidad.strip() or None
    jinete.celular = celular.strip() or None
    jinete.email = email.strip() or None
    jinete.provincia = provincia.strip() or None
    jinete.localidad = localidad.strip() or None
    jinete.categoria_habitual = categoria_habitual.strip() or None
    jinete.club_agrupacion = club_agrupacion.strip() or None
    jinete.observaciones = observaciones.strip() or None
    jinete.estado = estado

    db.commit()

    return RedirectResponse(
        url=f"/jinetes/{jinete.id}",
        status_code=303,
    )
